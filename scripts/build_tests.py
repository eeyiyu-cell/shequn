# -*- coding: utf-8 -*-
"""汇总三端测试用例 JSON → 专业测试用例 Excel（浅香槟风，与 PRD 呼应）。
读 scratchpad/tc/tc_adm.json + tc_pc.json + tc_h5.json，出 output/【测试用例】...xlsx。
"""
import os, json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TC = r"C:\Users\Admin\AppData\Local\Temp\claude\C--Users-Admin\5c8f5f13-1098-4076-9d24-28d84c193046\scratchpad\tc"
OUT = os.path.join(ROOT, "output", "【测试用例】红魔方社群积分系统-三端功能测试用例.xlsx")

CHAMP = "EAE0C8"; BRONZE = "5C4A2E"; BORDER = "D8CBA8"; ZEBRA = "FBF8F1"
PRI_COLOR = {"P0": "C0392B", "P1": "B9770E", "P2": "7A7A7A"}
FONT = "微软雅黑"

def load(name):
    p = os.path.join(TC, name)
    if not os.path.exists(p):
        print("缺文件:", p); return []
    with open(p, encoding="utf-8") as f:
        data = json.load(f)
    print(f"{name}: {len(data)} 条")
    return data

def as_text(v):
    if isinstance(v, list):
        return "\n".join(str(x) for x in v)
    return str(v)

cases = load("tc_adm.json") + load("tc_pc.json") + load("tc_h5.json")
if not cases:
    print("无用例，退出"); sys.exit(1)

thin = Side(style="thin", color=BORDER)
box = Border(left=thin, right=thin, top=thin, bottom=thin)
hfill = PatternFill("solid", fgColor=CHAMP)
zfill = PatternFill("solid", fgColor=ZEBRA)

wb = Workbook()

# ============ Sheet1 覆盖统计 ============
ws0 = wb.active; ws0.title = "覆盖统计"
ws0["A1"] = "【测试用例】红魔方社群积分系统 · 三端功能测试用例"
ws0["A1"].font = Font(name=FONT, size=16, bold=True, color="000000")
ws0["A2"] = "覆盖范围：PC 后台配置端 + 用户 PC web 门户 + 用户 H5 移动端；对应 PRD 三端合并版全部功能模块"
ws0["A2"].font = Font(name=FONT, size=10, color=BRONZE)

def stat_block(ws, start_row, title, pairs):
    ws.cell(start_row, 1, title).font = Font(name=FONT, size=12, bold=True, color=BRONZE)
    r = start_row + 1
    ws.cell(r, 1, "分类").font = Font(name=FONT, bold=True, color=BRONZE)
    ws.cell(r, 2, "用例数").font = Font(name=FONT, bold=True, color=BRONZE)
    ws.cell(r, 1).fill = hfill; ws.cell(r, 2).fill = hfill
    ws.cell(r, 1).border = box; ws.cell(r, 2).border = box
    r += 1
    for k, v in pairs:
        ws.cell(r, 1, k).font = Font(name=FONT); ws.cell(r, 2, v).font = Font(name=FONT)
        ws.cell(r, 1).border = box; ws.cell(r, 2).border = box
        r += 1
    return r + 1

def count_by(key):
    d = {}
    for c in cases:
        d[c.get(key, "")] = d.get(c.get(key, ""), 0) + 1
    return d

total = len(cases)
row = 4
row = stat_block(ws0, row, f"一、总计：{total} 条", [
    ("后台配置端", sum(1 for c in cases if c.get("end") == "后台配置端")),
    ("用户PC端", sum(1 for c in cases if c.get("end") == "用户PC端")),
    ("用户H5端", sum(1 for c in cases if c.get("end") == "用户H5端")),
    ("合计", total),
])
pri = count_by("priority")
row = stat_block(ws0, row, "二、优先级分布", [(k, pri.get(k, 0)) for k in ["P0", "P1", "P2"]])
typ = count_by("type")
row = stat_block(ws0, row, "三、用例类型分布", sorted(typ.items(), key=lambda x: -x[1]))

# 模块×用例数
ws0.cell(row, 1, "四、各功能模块用例数").font = Font(name=FONT, size=12, bold=True, color=BRONZE)
row += 1
for h, c in [("所属端", 1), ("功能模块", 2), ("用例数", 3)]:
    cell = ws0.cell(row, c, h); cell.font = Font(name=FONT, bold=True, color=BRONZE); cell.fill = hfill; cell.border = box
row += 1
seen = []
for c in cases:
    k = (c.get("end", ""), c.get("module", ""))
    if k not in seen: seen.append(k)
for end, mod in seen:
    n = sum(1 for c in cases if c.get("end") == end and c.get("module") == mod)
    ws0.cell(row, 1, end).font = Font(name=FONT); ws0.cell(row, 1).border = box
    ws0.cell(row, 2, mod).font = Font(name=FONT); ws0.cell(row, 2).border = box
    ws0.cell(row, 3, n).font = Font(name=FONT); ws0.cell(row, 3).border = box
    row += 1
ws0.column_dimensions["A"].width = 18; ws0.column_dimensions["B"].width = 30; ws0.column_dimensions["C"].width = 12

# ============ Sheet2 测试用例明细 ============
ws = wb.create_sheet("测试用例明细")
HEADERS = ["用例编号", "所属端", "功能模块", "用例标题", "优先级", "用例类型", "前置条件", "测试步骤", "预期结果"]
WIDTHS = [14, 12, 18, 40, 8, 10, 24, 50, 50]
for i, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
    cell = ws.cell(1, i, h)
    cell.font = Font(name=FONT, bold=True, color=BRONZE, size=11)
    cell.fill = hfill; cell.border = box
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 26

for ridx, c in enumerate(cases, start=2):
    vals = [c.get("id", ""), c.get("end", ""), c.get("module", ""), c.get("title", ""),
            c.get("priority", ""), c.get("type", ""), c.get("precondition", ""),
            as_text(c.get("steps", "")), as_text(c.get("expected", ""))]
    for i, v in enumerate(vals, 1):
        cell = ws.cell(ridx, i, v)
        cell.border = box
        cell.alignment = Alignment(vertical="top", wrap_text=True,
                                   horizontal="center" if i in (1, 2, 5, 6) else "left")
        if i == 5:  # 优先级配色
            cell.font = Font(name=FONT, bold=True, color=PRI_COLOR.get(str(v), BRONZE))
        else:
            cell.font = Font(name=FONT, color="262626")
        if ridx % 2 == 0:
            cell.fill = zfill

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(cases)+1}"

wb.save(OUT)
print("\nsaved:", OUT)
print(f"总用例 {total} 条；sheet: 覆盖统计 + 测试用例明细")
